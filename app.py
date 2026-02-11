"""DART main application window."""

import csv
import json
import os

import chardet
from PySide6.QtCore import QAbstractTableModel, QModelIndex, Qt, QTimer
from PySide6.QtGui import QAction, QFont, QKeySequence, QShortcut
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QTableView, QWidget, QVBoxLayout, QHBoxLayout,
    QLineEdit, QPushButton, QFileDialog, QMessageBox, QFrame, QMenu,
    QHeaderView, QLabel, QStackedWidget,
)

from theme import build_stylesheet, sys_font_family
from filter_model import FilteredTableModel
from dialogs import AboutDialog, HelpDialog, StatsDialog
from widgets import ColumnFilterBar, EmptyState, LoadingOverlay, SearchBar

# Optional Excel support
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

ASSETS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets")
LOGO_PATH = os.path.join(ASSETS_DIR, "logo.png")

# Supported file extensions
_OPEN_FILTER = "All Supported (*.csv *.tsv *.txt *.json *.xlsx);;CSV (*.csv);;TSV (*.tsv);;Text (*.txt);;JSON (*.json);;Excel (*.xlsx)"
_EXPORT_FILTER = "CSV (*.csv);;TSV (*.tsv);;JSON (*.json);;Excel (*.xlsx)"

_SUPPORTED_EXT = {".csv", ".tsv", ".txt", ".json", ".xlsx"}


# -------------------------------------------------------------------
# RawTableModel — stores data as plain Python lists
# -------------------------------------------------------------------

class RawTableModel(QAbstractTableModel):
    """Flat list-of-lists model. No QStandardItem overhead."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self._headers: list[str] = []
        self._data: list[list[str]] = []

    def load(self, headers: list[str], rows: list[list[str]]):
        self.beginResetModel()
        self._headers = headers
        self._data = rows
        self.endResetModel()

    # -- Direct access for workers --

    def raw_data(self) -> list[list[str]]:
        return self._data

    def raw_headers(self) -> list[str]:
        return self._headers

    # -- QAbstractTableModel interface --

    def rowCount(self, parent=QModelIndex()):
        if parent.isValid():
            return 0
        return len(self._data)

    def columnCount(self, parent=QModelIndex()):
        if parent.isValid():
            return 0
        if not self._headers:
            return 0
        return len(self._headers) + 1  # +1 for hidden index col 0

    def data(self, index, role=Qt.DisplayRole):
        if role != Qt.DisplayRole or not index.isValid():
            return None
        row = index.row()
        col = index.column()
        if col == 0:
            return row  # hidden index column
        data_col = col - 1
        if row < len(self._data) and data_col < len(self._data[row]):
            return self._data[row][data_col]
        return None

    def headerData(self, section, orientation, role=Qt.DisplayRole):
        if role != Qt.DisplayRole or orientation != Qt.Horizontal:
            return None
        if section == 0:
            return "_index_"
        idx = section - 1
        if idx < len(self._headers):
            return self._headers[idx]
        return None


# -------------------------------------------------------------------
# Main Window
# -------------------------------------------------------------------

class CSVFilterSortApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("DART")
        self.resize(1200, 720)
        self.theme_name = "dark"
        self.filter_editors: dict[int, QLineEdit] = {}
        self._init_models()
        self._init_ui()
        self.apply_theme()
        self.default_font_size = self.table_view.font().pixelSize()
        if self.default_font_size <= 0:
            self.default_font_size = self.table_view.font().pointSize()
        self._apply_table_density(self.default_font_size)
        self._update_zoom_label()

    # -----------------------------------------------------------------
    # Models
    # -----------------------------------------------------------------

    def _init_models(self):
        self.model = RawTableModel(self)
        self.proxy_model = FilteredTableModel(self)
        self.proxy_model.set_source(self.model)
        self.proxy_model.modelReset.connect(self._update_row_count)

    # -----------------------------------------------------------------
    # UI layout
    # -----------------------------------------------------------------

    def _init_ui(self):
        central = QWidget()
        central.setAttribute(Qt.WA_StyledBackground, True)
        self.setCentralWidget(central)
        root = QVBoxLayout(central)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        self.action_bar = self._build_action_bar()
        root.addWidget(self.action_bar)

        self.search_bar = SearchBar()
        self.search_bar.debounced_text.connect(
            lambda t: self.proxy_model.setGlobalFilter(t))
        root.addWidget(self.search_bar)

        self.stack = QStackedWidget()
        self.stack.setObjectName("contentStack")
        self.stack.setAttribute(Qt.WA_StyledBackground, True)
        root.addWidget(self.stack, 1)

        self.empty_state = EmptyState(LOGO_PATH, self.open_file)
        self.stack.addWidget(self.empty_state)

        data_page = QWidget()
        data_page.setObjectName("dataPage")
        data_page.setAttribute(Qt.WA_StyledBackground, True)
        data_layout = QVBoxLayout(data_page)
        data_layout.setContentsMargins(0, 0, 0, 0)
        data_layout.setSpacing(0)

        self.table_view = self._build_table()
        self.column_filter_bar = ColumnFilterBar(self.table_view)
        data_layout.addWidget(self.column_filter_bar)
        data_layout.addWidget(self.table_view, 1)

        self.stack.addWidget(data_page)
        self.stack.setCurrentIndex(0)

        self.loading_overlay = LoadingOverlay(LOGO_PATH, self.stack)

        self._setup_status_bar()
        self._build_menus()
        self._bind_shortcuts()
        self.setAcceptDrops(True)

    def _build_table(self) -> QTableView:
        tv = QTableView()
        tv.setSortingEnabled(True)
        tv.setAlternatingRowColors(True)
        tv.setShowGrid(False)
        tv.verticalHeader().setVisible(False)
        tv.verticalHeader().setDefaultSectionSize(36)
        tv.setContextMenuPolicy(Qt.CustomContextMenu)
        tv.customContextMenuRequested.connect(self._context_menu)
        hdr = tv.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.Interactive)
        hdr.setHighlightSections(False)
        hdr.setStretchLastSection(False)
        hdr.setMinimumSectionSize(60)
        tv.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        tv.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        tv.setModel(self.proxy_model)
        return tv

    # -----------------------------------------------------------------
    # Action bar
    # -----------------------------------------------------------------

    def _build_action_bar(self):
        bar = QWidget()
        bar.setObjectName("actionBar")
        bar.setAttribute(Qt.WA_StyledBackground, True)
        bar.setFixedHeight(40)
        lay = QHBoxLayout(bar)
        lay.setContentsMargins(8, 0, 8, 0)
        lay.setSpacing(2)

        def btn(text, tip, cb):
            b = QPushButton(text)
            b.setObjectName("actionButton")
            b.setToolTip(tip)
            b.setCursor(Qt.PointingHandCursor)
            b.clicked.connect(cb)
            return b

        def sep():
            s = QFrame()
            s.setObjectName("toolSep")
            s.setFrameShape(QFrame.VLine)
            s.setFixedWidth(1)
            s.setFixedHeight(20)
            return s

        lay.addWidget(btn("Open", "Open file  (Ctrl+O)", self.open_file))
        lay.addWidget(btn("Export", "Export data  (Ctrl+Shift+E)", self.export_data))
        lay.addWidget(sep())
        lay.addWidget(btn("Clear Filters", "Clear all filters", self.clear_filters))
        lay.addWidget(btn("Reset Sort", "Reset to original order", self.reset_sorting))
        lay.addWidget(btn(
            "Fit Columns",
            "Auto-resize columns (balanced). Shift-click for exact full fit.",
            self.fit_columns,
        ))
        lay.addWidget(sep())
        lay.addWidget(btn("Stats", "Column statistics  (Ctrl+I)", self.show_stats))
        lay.addWidget(sep())
        lay.addWidget(btn("-", "Zoom out  (Ctrl+-)", self.zoom_out))
        lay.addWidget(btn("+", "Zoom in  (Ctrl+=)", self.zoom_in))
        lay.addWidget(sep())
        lay.addStretch()
        lay.addWidget(btn("Save", "Save settings", self.save_settings))
        lay.addWidget(btn("Load", "Load settings", self.load_settings))
        lay.addWidget(sep())
        self._theme_btn = btn("Light", "Toggle theme  (Ctrl+Shift+T)", self.toggle_theme)
        lay.addWidget(self._theme_btn)
        lay.addWidget(btn("Help", "Help  (F1)", self.show_help))
        return bar

    # -----------------------------------------------------------------
    # Menus & shortcuts
    # -----------------------------------------------------------------

    def _build_menus(self):
        mb = self.menuBar()
        fm = mb.addMenu("&File")
        self._menu_action(fm, "Open...", self.open_file, "Ctrl+O")
        self._menu_action(fm, "Export...", self.export_data, "Ctrl+Shift+E")
        fm.addSeparator()
        self._menu_action(fm, "Save Settings", self.save_settings)
        self._menu_action(fm, "Load Settings", self.load_settings)
        fm.addSeparator()
        self._menu_action(fm, "Exit", self.close, "Ctrl+Q")

        hm = mb.addMenu("&Help")
        self._menu_action(hm, "Help", self.show_help, "F1")
        self._menu_action(hm, "About", self.show_about)

    def _menu_action(self, menu, text, slot, shortcut=None):
        a = QAction(text, self)
        a.triggered.connect(slot)
        if shortcut:
            a.setShortcut(shortcut)
        menu.addAction(a)

    def _bind_shortcuts(self):
        QShortcut(QKeySequence("Ctrl+F"), self).activated.connect(
            lambda: self.search_bar.input.setFocus())
        QShortcut(QKeySequence("Ctrl+="), self).activated.connect(self.zoom_in)
        QShortcut(QKeySequence("Ctrl+-"), self).activated.connect(self.zoom_out)
        QShortcut(QKeySequence("Ctrl+0"), self).activated.connect(self.reset_zoom)
        QShortcut(QKeySequence("Ctrl+R"), self).activated.connect(self.fit_columns)
        QShortcut(QKeySequence("Ctrl+Shift+R"), self).activated.connect(
            lambda: self.resize_columns("full"))
        QShortcut(QKeySequence("Ctrl+I"), self).activated.connect(self.show_stats)
        QShortcut(QKeySequence("Ctrl+Shift+T"), self).activated.connect(self.toggle_theme)

    # -----------------------------------------------------------------
    # Status bar
    # -----------------------------------------------------------------

    def _setup_status_bar(self):
        self.status_file = QLabel("No file loaded")
        self.status_rows = QLabel("")
        self.status_zoom = QLabel("100%")
        bar = self.statusBar()
        bar.addWidget(self.status_file, 1)
        bar.addPermanentWidget(self.status_rows)
        bar.addPermanentWidget(self.status_zoom)

    def _update_row_count(self):
        total = self.model.rowCount()
        visible = self.proxy_model.rowCount()
        if total:
            self.status_rows.setText(f"{visible:,} of {total:,} rows")
        else:
            self.status_rows.setText("")

    def _update_zoom_label(self):
        pct = round(self._get_font_size() / self.default_font_size * 100)
        self.status_zoom.setText(f"{pct}%")

    # -----------------------------------------------------------------
    # Theme
    # -----------------------------------------------------------------

    def apply_theme(self):
        self.setStyleSheet(build_stylesheet(self.theme_name))
        self._theme_btn.setText("Light" if self.theme_name == "dark" else "Dark")
        from widgets import FilterInput
        if FilterInput._popup is not None:
            FilterInput._popup._apply_style(self.theme_name)

    def toggle_theme(self):
        self.theme_name = "light" if self.theme_name == "dark" else "dark"
        self.apply_theme()

    # -----------------------------------------------------------------
    # Open file (multi-format)
    # -----------------------------------------------------------------

    def open_file(self):
        path, _ = QFileDialog.getOpenFileName(self, "Open File", "", _OPEN_FILTER)
        if path:
            self._load_file(path)

    def _load_file(self, path):
        ext = os.path.splitext(path)[1].lower()
        self.loading_overlay.resize(self.stack.size())
        self.loading_overlay.show_overlay("Loading...")
        QApplication.processEvents()

        try:
            if ext == ".xlsx":
                headers, rows = self._read_excel(path)
            elif ext == ".json":
                headers, rows = self._read_json(path)
            else:
                headers, rows = self._read_delimited(path, ext)

            if not headers:
                self.loading_overlay.hide_overlay()
                QMessageBox.warning(self, "Empty File", "No data found.")
                return

            self._populate_model(headers, rows)
            self.status_file.setText(os.path.basename(path))
            self._update_row_count()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Could not load file:\n{e}")
        finally:
            self.loading_overlay.hide_overlay()

    # --- Readers ---

    def _read_delimited(self, path, ext):
        """Read CSV, TSV, or other delimited text. Auto-detects delimiter."""
        with open(path, "rb") as f:
            raw = f.read(10000)
        detected = chardet.detect(raw)
        encoding = detected.get("encoding", "utf-8")
        if encoding and encoding.lower() == "ascii":
            encoding = "latin1"

        with open(path, newline="", encoding=encoding) as fh:
            sample = fh.read(8192)

        # Determine delimiter
        if ext == ".tsv":
            delim = "\t"
        else:
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
                delim = dialect.delimiter
            except csv.Error:
                delim = ","

        with open(path, newline="", encoding=encoding) as fh:
            reader = csv.reader(fh, delimiter=delim)
            data = list(reader)

        if not data:
            return [], []
        return data[0], data[1:]

    def _read_excel(self, path):
        """Read .xlsx via openpyxl."""
        if not HAS_OPENPYXL:
            raise ImportError(
                "Excel support requires openpyxl.\n"
                "Install it with:  pip install openpyxl")
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = [str(c) if c is not None else "" for c in next(rows_iter)]
        rows = []
        for row in rows_iter:
            rows.append([str(c) if c is not None else "" for c in row])
        wb.close()
        return headers, rows

    def _read_json(self, path):
        """Read JSON - expects array of objects (records)."""
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, dict):
            # If top-level is a dict, look for first array value
            for v in data.values():
                if isinstance(v, list):
                    data = v
                    break
            else:
                # Wrap single object
                data = [data]
        if not isinstance(data, list) or not data:
            return [], []
        # Collect all keys across all records for headers
        headers = []
        seen = set()
        for record in data:
            if isinstance(record, dict):
                for k in record:
                    if k not in seen:
                        headers.append(str(k))
                        seen.add(k)
        rows = []
        for record in data:
            if isinstance(record, dict):
                rows.append([str(record.get(h, "")) for h in headers])
        return headers, rows

    # --- Populate model ---

    def _populate_model(self, headers, rows):
        self.model.load(headers, rows)
        self.proxy_model.reset_all()
        self.table_view.setColumnHidden(0, True)
        self._setup_column_filters(headers)
        self.resize_columns()
        self.table_view.reset()
        self.stack.setCurrentIndex(1)

    def _setup_column_filters(self, headers):
        self.filter_editors.clear()
        self.column_filter_bar.rebuild(headers)
        for i, inp in enumerate(self.column_filter_bar.inputs):
            col = i + 1
            inp.debounced_text.connect(
                lambda text, c=col: self._on_filter(c, text))
            self.filter_editors[col] = inp

    def _on_filter(self, col, text):
        self.proxy_model.setFilterForColumn(col, text)

    # -----------------------------------------------------------------
    # Filters & sorting
    # -----------------------------------------------------------------

    def clear_filters(self):
        self.column_filter_bar.clear_all()
        self.search_bar.input.clear()
        self._update_row_count()

    def reset_sorting(self):
        self.proxy_model.sort(0, Qt.AscendingOrder)

    def fit_columns(self):
        mods = QApplication.keyboardModifiers()
        if mods & Qt.ShiftModifier:
            mode = "full"
        elif mods & Qt.AltModifier:
            mode = "quick"
        else:
            mode = "balanced"
        self.resize_columns(mode)

    def _visible_proxy_rows(self):
        total = self.proxy_model.rowCount()
        if total <= 0:
            return []
        first = self.table_view.rowAt(0)
        if first < 0:
            first = 0
        last = self.table_view.rowAt(max(0, self.table_view.viewport().height() - 1))
        if last < 0:
            last = min(total - 1, first + 40)
        first = max(0, min(first, total - 1))
        last = max(first, min(last, total - 1))
        return list(range(first, last + 1))

    def _sample_proxy_rows(self, mode):
        total = self.proxy_model.rowCount()
        if total <= 0:
            return []
        if mode == "full":
            return list(range(total))

        rows = set(self._visible_proxy_rows())
        if mode == "quick":
            return sorted(rows)

        # Balanced mode:
        # - Always include visible rows.
        # - Add head/tail windows.
        # - Add evenly spaced rows across the filtered dataset.
        edge = min(30, total)
        rows.update(range(edge))
        rows.update(range(max(0, total - edge), total))

        target = 240
        if total <= target:
            rows.update(range(total))
        else:
            step = max(1, total // target)
            for r in range(0, total, step):
                rows.add(r)
            rows.add(total - 1)
        return sorted(rows)

    def resize_columns(self, mode="balanced"):
        """Auto-fit columns.

        Modes:
        - quick: visible rows only (fastest)
        - balanced: visible + stratified sample (default)
        - full: all filtered rows (most accurate, slowest)
        """
        if self.model.rowCount() == 0 or self.model.columnCount() <= 1:
            return

        mode = (mode or "balanced").lower()
        if mode not in {"quick", "balanced", "full"}:
            mode = "balanced"

        proxy_rows = self._sample_proxy_rows(mode)
        if not proxy_rows:
            return

        header = self.table_view.horizontalHeader()
        fm = self.table_view.fontMetrics()
        col_range = range(1, self.model.columnCount())

        viewport_w = max(600, self.table_view.viewport().width())
        max_col_w = max(180, int(viewport_w * 0.55))
        min_col_w = 72
        header_pad = 30
        cell_pad = 20
        measure_limit = 180

        widths = {}
        for col in col_range:
            title = str(self.model.headerData(col, Qt.Horizontal) or "")
            widths[col] = max(min_col_w, fm.horizontalAdvance(title) + header_pad)

        raw_rows = self.model.raw_data()
        visible_indices = self.proxy_model._visible_indices
        capped = {col: False for col in col_range}

        QApplication.setOverrideCursor(Qt.WaitCursor)
        try:
            for i, proxy_row in enumerate(proxy_rows):
                if proxy_row < 0 or proxy_row >= len(visible_indices):
                    continue
                src_row = visible_indices[proxy_row]
                if src_row < 0 or src_row >= len(raw_rows):
                    continue
                row_data = raw_rows[src_row]

                for col in col_range:
                    if capped[col]:
                        continue
                    data_col = col - 1
                    if data_col >= len(row_data):
                        continue
                    txt = row_data[data_col]
                    if not txt:
                        continue
                    if len(txt) > measure_limit:
                        txt = txt[:measure_limit]
                    w = fm.horizontalAdvance(txt) + cell_pad
                    if w > widths[col]:
                        if w >= max_col_w:
                            widths[col] = max_col_w
                            capped[col] = True
                        else:
                            widths[col] = w

                if mode == "full" and i % 25000 == 0:
                    QApplication.processEvents()
        finally:
            QApplication.restoreOverrideCursor()

        for col in col_range:
            header.resizeSection(col, int(max(min_col_w, min(widths[col], max_col_w))))

        if mode == "full":
            detail = "all filtered rows"
        elif mode == "quick":
            detail = "visible rows"
        else:
            detail = f"{len(proxy_rows):,} sampled rows"
        self.statusBar().showMessage(f"Fit columns ({mode}): {detail}", 3000)

    # -----------------------------------------------------------------
    # Export (multi-format) — direct list access
    # -----------------------------------------------------------------

    def export_data(self):
        if self.model.rowCount() == 0:
            QMessageBox.information(self, "No Data", "No data to export.")
            return
        path, chosen = QFileDialog.getSaveFileName(
            self, "Export Data", "", _EXPORT_FILTER)
        if not path:
            return

        ext = os.path.splitext(path)[1].lower()
        headers = list(self.model.raw_headers())

        def visible_rows():
            data = self.model.raw_data()
            for src_row in self.proxy_model._visible_indices:
                yield data[src_row]

        try:
            if ext == ".json":
                self._export_json(path, headers, visible_rows)
            elif ext == ".xlsx":
                self._export_excel(path, headers, visible_rows)
            elif ext == ".tsv":
                self._export_delimited(path, headers, visible_rows, "\t")
            else:
                self._export_delimited(path, headers, visible_rows, ",")
            self.statusBar().showMessage(f"Exported to {path}", 3000)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Could not export:\n{e}")

    def _export_delimited(self, path, headers, rows_fn, delimiter):
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f, delimiter=delimiter)
            w.writerow(headers)
            for row in rows_fn():
                w.writerow(row)

    def _export_json(self, path, headers, rows_fn):
        records = []
        for row in rows_fn():
            records.append(dict(zip(headers, row)))
        with open(path, "w", encoding="utf-8") as f:
            json.dump(records, f, indent=2, ensure_ascii=False)

    def _export_excel(self, path, headers, rows_fn):
        if not HAS_OPENPYXL:
            raise ImportError(
                "Excel export requires openpyxl.\n"
                "Install it with:  pip install openpyxl")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "DART Export"
        ws.append(headers)
        for row in rows_fn():
            ws.append(row)
        wb.save(path)

    # -----------------------------------------------------------------
    # Settings
    # -----------------------------------------------------------------

    def save_settings(self):
        if self.model.columnCount() <= 1:
            QMessageBox.information(self, "No Data", "Load a file first.")
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Save Settings", "", "JSON Files (*.json)")
        if path:
            try:
                hdr = self.table_view.horizontalHeader()
                sort_order = hdr.sortIndicatorOrder()
                if hasattr(sort_order, "value"):
                    sort_order = sort_order.value
                data = {
                    "filters": {str(c): e.text()
                                for c, e in self.filter_editors.items()},
                    "sort": {"section": hdr.sortIndicatorSection(),
                             "order": int(sort_order)},
                }
                with open(path, "w", encoding="utf-8") as f:
                    json.dump(data, f, indent=4)
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Could not save:\n{e}")

    def load_settings(self):
        if self.model.columnCount() <= 1:
            QMessageBox.information(self, "No Data", "Load a file first.")
            return
        path, _ = QFileDialog.getOpenFileName(
            self, "Load Settings", "", "JSON Files (*.json)")
        if path:
            try:
                with open(path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                for col_str, text in data.get("filters", {}).items():
                    col = int(col_str)
                    if col in self.filter_editors:
                        self.filter_editors[col].setText(text)
                s = data.get("sort", {})
                sec = s.get("section", -1)
                raw_order = s.get("order", 0)
                if hasattr(raw_order, "value"):
                    raw_order = raw_order.value
                try:
                    sort_order = Qt.SortOrder(int(raw_order))
                except (TypeError, ValueError):
                    sort_order = Qt.AscendingOrder
                if 0 <= sec < self.model.columnCount():
                    self.proxy_model.sort(sec, sort_order)
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Could not load:\n{e}")

    # -----------------------------------------------------------------
    # Zoom
    # -----------------------------------------------------------------

    def _get_font_size(self):
        f = self.table_view.font()
        s = f.pixelSize()
        return s if s > 0 else f.pointSize()

    def _set_font_size(self, size):
        f = self.table_view.font()
        if f.pixelSize() > 0:
            f.setPixelSize(size)
        else:
            f.setPointSize(size)
        self.table_view.setFont(f)
        self._apply_table_density(size)
        self._update_zoom_label()

    def _apply_table_density(self, font_size):
        """Scale row height with zoom so zoom-out shows more rows."""
        row_height = max(16, int(round(font_size * 2.75)))
        self.table_view.verticalHeader().setDefaultSectionSize(row_height)

    def zoom_in(self):
        self._set_font_size(self._get_font_size() + 1)

    def zoom_out(self):
        self._set_font_size(max(6, self._get_font_size() - 1))

    def reset_zoom(self):
        self._set_font_size(self.default_font_size)

    # -----------------------------------------------------------------
    # Dialogs
    # -----------------------------------------------------------------

    def show_help(self):
        HelpDialog(self.theme_name, self).exec()

    def show_about(self):
        AboutDialog(self).exec()

    def show_stats(self):
        StatsDialog(self.model, self.proxy_model, self).exec()

    # -----------------------------------------------------------------
    # Context menu
    # -----------------------------------------------------------------

    def _context_menu(self, pos):
        idx = self.table_view.indexAt(pos)
        if not idx.isValid():
            return
        menu = QMenu()
        copy = QAction("Copy Cell", self)
        copy.triggered.connect(
            lambda: QApplication.clipboard().setText(
                str(self.proxy_model.data(idx, Qt.DisplayRole))))
        menu.addAction(copy)
        menu.exec(self.table_view.viewport().mapToGlobal(pos))

    # -----------------------------------------------------------------
    # Drag & drop (all supported formats)
    # -----------------------------------------------------------------

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            for url in event.mimeData().urls():
                ext = os.path.splitext(url.toLocalFile())[1].lower()
                if ext in _SUPPORTED_EXT:
                    event.acceptProposedAction()
                    return

    def dropEvent(self, event):
        for url in event.mimeData().urls():
            path = url.toLocalFile()
            ext = os.path.splitext(path)[1].lower()
            if ext in _SUPPORTED_EXT:
                self._load_file(path)
                break

    # -----------------------------------------------------------------
    # Resize overlay
    # -----------------------------------------------------------------

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self.loading_overlay.resize(self.stack.size())

    # -----------------------------------------------------------------
    # Close — clean up worker thread
    # -----------------------------------------------------------------

    def closeEvent(self, event):
        self.proxy_model.shutdown()
        super().closeEvent(event)
